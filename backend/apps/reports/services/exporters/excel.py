"""openpyxl-based Excel export — see
apps/reports/services/exporters/pdf.py's docstring for why this is one
generic table renderer rather than one function per report type.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def render_table_xlsx(*, title, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Report"  # Excel's sheet-name length limit

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row in rows:
        sheet.append([str(cell) for cell in row])

    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = max((len(value) for value in values), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(10, width + 2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
